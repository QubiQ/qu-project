# Copyright 2018 Valentín Vinagre <valentin.vinagre@qubiq.es>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

from odoo import models, fields, tools, _
from io import BytesIO
import codecs
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Border, Side, Font
from datetime import datetime
import os


class ReportHoursConsumption(models.TransientModel):
    _inherit = 'file.download.model'
    _name = "report.hours.consumption"

    contract_id = fields.Many2one(
        comodel_name='project.contract',
        string=_('Project'),
        required=True,
        ondelete='set null'
    )
    type_hours = fields.Selection(
        string=_('Type of hours'),
        required=True,
        default='all',
        selection=[
            ('billable', 'Billable'),
            ('nobillable', 'No Billable'),
            ('all', 'All')
        ]
    )

    def get_filename(self):
        name = _('Hours Consumption')
        return name + '.xlsx'

    def _set_styles(self, wb):
        general_style = NamedStyle(
            name='general_style'
        )
        general_style.font = Font(
            name='Arial',
            bold=False,
            size=8,
            color='000000'
        )
        general_style.alignment = Alignment(
            horizontal='left'
        )
        hours_style = NamedStyle(
            name="hours_style"
        )
        hours_style.font = Font(
            name='Arial',
            bold=False,
            size=8,
            color='000000'
        )
        hours_style.alignment = Alignment(
            horizontal='right'
        )
        bd = Side(
            style='thin',
            color='000000'
        )
        start_table_style = NamedStyle(
            name='start_table_style'
        )
        start_table_style.font = Font(
            name='Arial',
            bold=False,
            size=11,
            color='000000'
        )
        start_table_style.alignment = Alignment(
            horizontal='left'
        )
        start_table_style.border = Border(
            left=bd,
            top=bd,
            bottom=bd
        )
        mid_table_style = NamedStyle(
            name='mid_table_style'
        )
        mid_table_style.font = Font(
            name='Arial',
            bold=False,
            size=11,
            color='000000'
        )
        mid_table_style.alignment = Alignment(
            horizontal='right'
        )
        mid_table_style.border = Border(
            top=bd,
            bottom=bd
        )
        end_table_style = NamedStyle(
            name='end_table_style'
        )
        end_table_style.font = Font(
            name='Arial',
            bold=False,
            size=11,
            color='000000'
        )
        end_table_style.alignment = Alignment(
            horizontal='right'
        )
        end_table_style.border = Border(
            right=bd,
            top=bd,
            bottom=bd
        )
        wb.add_named_style(general_style)
        wb.add_named_style(hours_style)
        wb.add_named_style(start_table_style)
        wb.add_named_style(mid_table_style)
        wb.add_named_style(end_table_style)

    """
    Returns a string date given a datetime/date date
    """
    def _format_date(self, date):
        res = ''
        if date:
            res = str(date.day).zfill(2) + '/' + str(date.month).zfill(2) +\
                '/' + str(date.year).zfill(2)
        return res

    def get_type(self, aal):
        if aal.contract_hours:
            kind = 'Bolsa'
        elif not aal.move_id and aal.task_id:
            kind = 'Tarea'
        else:
            kind = 'Manual'
        return kind

    def get_creation_date(self, aal, kind):
        return ''

    def get_contact(self, aal, kind):
        return ''

    def get_origin(self, aal, kind):
        return ''

    def get_state(self, aal, kind):
        return ''

    def get_content(self):
        # Path creation to find the excel
        excel_path = ''
        for ad in tools.config['addons_path'].split(','):
            if os.path.exists(ad+'/report_hours_consumption/static/template'):
                excel_path = ad+'/report_hours_consumption/static/template/' +\
                    'base.xlsx'
        # The path is instantiated
        path = os.path.expanduser(excel_path)
        wb = load_workbook(path)
        # The first page of the excel is instantiated
        ws = wb.worksheets[0]
        # Total of letters in the excel
        letters_list = [
            'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
            'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
            'W', 'X', 'Y', 'Z'
        ]
        fix_col = 12
        # Save the number of columns we will fill
        col_list = letters_list[:fix_col]
        line = 5
        # Styles
        self._set_styles(wb)
        # Company logo
        if self.env.user.company_id.logo:
            image_stream = BytesIO(codecs.decode(
                self.env.user.company_id.logo, 'base64'))
            image = Image(image_stream)
            baseheight = 63.87
            hpercent = (baseheight/float(image.height))
            wsize = int((float(image.width)*float(hpercent)))
            image.width = wsize
            image.height = baseheight
            ws.add_image(image, 'I2')
        # Contract name
        ws['B1'] = self.contract_id.name
        # Date
        ws['J1'] = self._format_date(datetime.now())
        ws['J1'].style = 'general_style'
        # Body
        domain = [('contract_id', '=', self.contract_id.id)]
        if self.type_hours != 'all':
            domain.append(('billable', '=', self.type_hours == 'billable'))
        for aal in self.env['account.analytic.line'].search(domain).filtered(
           lambda x:
           (not x.product_id) or
           (x.product_id and x.product_id.categ_id.show_hours_report)
        ).sorted(key='create_date'):
            ws.row_dimensions[line].height = 20
            kind = self.get_type(aal)
            # Type
            col = 0
            kind_name = kind
            if kind == 'Tarea':
                kind_name = str(aal.project_id.label_tasks)
            ws[col_list[col]+str(line)] = kind_name
            ws[col_list[col]+str(line)].style = 'general_style'
            # Creation date
            creation_date = ''
            if kind == 'Bolsa':
                creation_date = self._format_date(
                    fields.Datetime.from_string(
                        aal.move_id.invoice_id.date_invoice
                    )
                )
            elif kind == 'Tarea':
                creation_date = self._format_date(
                    fields.Datetime.from_string(
                        aal.task_id.create_date
                    )
                )
            else:
                creation_date = self._format_date(
                    fields.Datetime.from_string(
                        self.get_creation_date(aal, kind)
                    )
                )
            col += 1
            ws[col_list[col]+str(line)] = creation_date
            ws[col_list[col]+str(line)].style = 'general_style'
            # Contact
            col += 1
            contact = ''
            if kind == 'Bolsa':
                contact = aal.move_id.invoice_id.partner_id.name or ''
            elif kind == 'Tarea':
                contact = aal.task_id.partner_id.name or ''
            else:
                contact = self.get_contact(aal, kind)
            ws[col_list[col]+str(line)] = contact
            ws[col_list[col]+str(line)].style = 'general_style'
            # Origin
            col += 1
            origin = ''
            if kind == 'Bolsa':
                origin = aal.move_id.invoice_id.number or ''
            elif kind == 'Tarea':
                origin = aal.task_id.name
            else:
                origin = self.get_origin(aal, kind)
            ws[col_list[col]+str(line)] = origin
            ws[col_list[col]+str(line)].style = 'general_style'
            # State
            col += 1
            state = ''
            if kind == 'Bolsa':
                state = aal.move_id.invoice_id.state or ''
            elif kind == 'Tarea':
                state = aal.task_id.stage_id.name or ''
            else:
                state = self.get_state(aal, kind)
            ws[col_list[col]+str(line)] = state
            ws[col_list[col]+str(line)].style = 'general_style'
            # Register date
            col += 1
            ws[col_list[col]+str(line)] =\
                self._format_date(fields.Datetime.from_string(aal.create_date))
            ws[col_list[col]+str(line)].style = 'general_style'
            # Resource
            col += 1
            resource = ''
            if kind == 'Bolsa':
                resource =\
                    aal.move_id.invoice_id.user_id.employee_ids[0].name or ''
            else:
                resource = aal.employee_id.name or ''
            ws[col_list[col]+str(line)] = resource
            ws[col_list[col]+str(line)].style = 'general_style'
            # Description
            col += 1
            ws[col_list[col]+str(line)] = aal.name
            ws[col_list[col]+str(line)].style = 'general_style'
            # Consumed hours
            col += 1
            ws[col_list[col]+str(line)] =\
                round(aal.unit_amount, 2) if kind != 'Bolsa' else 0.0
            ws[col_list[col]+str(line)].style = 'hours_style'
            # Bought hours
            col += 1
            ws[col_list[col]+str(line)] =\
                round(aal.unit_amount, 2) if kind == 'Bolsa' else 0.0
            ws[col_list[col]+str(line)].style = 'hours_style'
            # Scope
            col += 1
            scope = ''
            if kind == 'Tarea' and aal.task_id.beyond_scope:
                scope = 'Sí'
            else:
                scope = 'No'
            ws[col_list[col]+str(line)] = scope
            ws[col_list[col]+str(line)].style = 'general_style'
            # Billable
            col += 1
            billable = ''
            if aal.billable:
                billable = 'Sí'
            else:
                billable = 'No'
            ws[col_list[col]+str(line)] = billable
            ws[col_list[col]+str(line)].style = 'general_style'
            line += 1
        ws.row_dimensions[line].height = 30
        ws.row_dimensions[line+2].height = 30
        ws['H'+str(line)] = 'Total'
        ws['H'+str(line)].style = 'start_table_style'
        ws['I'+str(line)] = '=SUM(I%s:I%s)' % (5, str(line-1))
        ws['I'+str(line)].style = 'mid_table_style'
        ws['J'+str(line)] = '=SUM(J%s:J%s)' % (5, str(line-1))
        ws['J'+str(line)].style = 'end_table_style'
        line += 2
        ws['H'+str(line)] = 'Diferencia'
        ws['H'+str(line)].style = 'start_table_style'
        ws['I'+str(line)] = '=SUM(J%s-I%s)' % (str(line-2), str(line-2))
        ws['I'+str(line)].style = 'end_table_style'
        # Returns the excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
